import pandas as pd
import openpyxl
import os
from openpyxl.styles import Alignment, Font  # 修改：新增导入 Font 模块
from hk_calendar import get_real_date_from_sheet_name

def update_summary_sheet(excel_file, current_sheet_name):
    if not os.path.exists(excel_file):
        print(f"找不到文件 {excel_file}")
        return

    print(f"\n将 '{current_sheet_name}' 填入 summary 表...")

    try:
        # 1. 用 Pandas 读取数据
        df_daily = pd.read_excel(excel_file, sheet_name=current_sheet_name)
        id_col = df_daily.columns[0]  
        name_col = df_daily.columns[1] 
        share_col = next((c for c in df_daily.columns if "持股" in str(c)), df_daily.columns[-2])

        df_daily[id_col] = df_daily[id_col].astype(str).str.strip()
        df_daily[share_col] = df_daily[share_col].astype(str).str.replace(',', '', regex=True)
        df_daily[share_col] = pd.to_numeric(df_daily[share_col], errors='coerce')

        today_data = dict(zip(df_daily[id_col], df_daily[share_col]))
        today_names = dict(zip(df_daily[id_col], df_daily[name_col]))

        # 2. 使用 openpyxl 操作
        wb = openpyxl.load_workbook(excel_file)
        ws = wb['summary']

        # 3. 智能寻找最右侧的非空列
        # 从右向左查找，直到找到有内容的列
        max_col = ws.max_column
        target_col = max_col + 1
        
        # 智能查找：从右向左检查第2行（表头行），找到最后一个有内容的列
        for col in range(max_col, 1, -1):
            if ws.cell(row=2, column=col).value:
                target_col = col + 1
                break
        else:
            # 如果所有列都为空，默认从第4列开始
            target_col = 4
        
        # --- 设置新列：持股量 ---
        target_header_cell = ws.cell(row=2, column=target_col, value=f"{current_sheet_name}持股量")
        
        # ✨ 新增：设置加粗 + 自动换行 + 居中
        target_header_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        target_header_cell.font = Font(bold=True)  # <-- 加粗字体

        # 4. 扫描旧表 ID
        existing_ids = {str(ws.cell(row=r, column=2).value).strip(): r 
                        for r in range(3, ws.max_row + 1) if ws.cell(row=r, column=2).value}

        # 5. 填入数据
        next_new_row = ws.max_row + 1
        for comp_id, shares in today_data.items():
            if comp_id in existing_ids:
                ws.cell(row=existing_ids[comp_id], column=target_col, value=shares)
            else:
                ws.cell(row=next_new_row, column=2, value=comp_id)
                ws.cell(row=next_new_row, column=3, value=today_names.get(comp_id, ""))
                ws.cell(row=next_new_row, column=target_col, value=shares)
                next_new_row += 1

        # 6. 添加差值列
        if target_col >= 4:
            prev_col = target_col - 2
            prev_header = ws.cell(row=2, column=prev_col).value
            
            import re
            def extract_date(header):
                if isinstance(header, str):
                    date_match = re.search(r'\d+月\d+日', header)
                    return date_match.group(0) if date_match else header
                return str(header)
            
            current_date = extract_date(current_sheet_name)
            prev_date = extract_date(prev_header)
            
            # --- 设置新列：差值 ---
            diff_header = f"{current_date}\nVS{prev_date}"
            diff_col = target_col + 1
            diff_header_cell = ws.cell(row=2, column=diff_col, value=diff_header)
            
            # ✨ 新增：设置加粗 + 自动换行 + 居中
            diff_header_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            diff_header_cell.font = Font(bold=True)  # <-- 加粗字体
            
            # 填充差值
            for r in range(3, next_new_row):
                curr_v = ws.cell(row=r, column=target_col).value
                prev_v = ws.cell(row=r, column=prev_col).value
                if isinstance(curr_v, (int, float)) and isinstance(prev_v, (int, float)):
                    ws.cell(row=r, column=diff_col, value=curr_v - prev_v)

        # 7. 在持股量列上方添加真实日期列
        real_date = get_real_date_from_sheet_name(current_sheet_name)
        if real_date:
            # 真实日期列在持股量列的上方（第1行）
            real_date_cell = ws.cell(row=1, column=target_col, value=real_date)
            # 设置样式：加粗 + 红色 + 自动换行 + 居中
            real_date_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
            real_date_cell.font = Font(bold=True, color='FF0000')  # 红色字体
            print(f"已添加真实日期: {real_date}")

        # 8. 保存
        wb.save(excel_file)
        # 9. 关闭工作簿
        wb.close()
        print(f"成功！新列已加粗并自动换行。")

    except PermissionError:
        print(f"失败：请先关闭 Excel 文件！")
    except Exception as e:
        print(f"错误: {e}")